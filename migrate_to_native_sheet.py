"""
One-off migration: copy every tab, every cell value, every formula from the
old Office-compat xlsx workbook into Ahmad's new native Google Sheet. Run
once, then sheets_tools.py switches to reading/writing the new sheet
directly via the Sheets API (no more download/local-edit/human-import).
"""

import asyncio
import io

import openpyxl

import sheets_tools

OLD_SPREADSHEET_ID = "1pYMxHUX5oH-DD1VPSH2z3kGySLvVN3ce"
NEW_SPREADSHEET_ID = "1VvQpaYSUF668MQjr-w7VYwiThQozXmR3zdruw5C_V08"

TAB_ORDER = [
    "Accounts Overview", "Instructions", "Daily Production List", "Winner Tracking",
    "Link Funnel", "Scaling Log", "Trial Reel Waves", "Target Creator List",
]

# The one real Scaling Log entry that only ever existed locally (Jerome was
# already messaged about it, but Ahmad never clicked "import" on the old
# xlsx flow, so it never actually landed in a sheet until now).
MISSING_SCALING_LOG_ROW = [
    "2026-08-05", "lunaxvale", "https://www.instagram.com/reel/DbjcfCTh1ta/?igsh=MWk2eTFlNXRmeHNmNg==",
    "Debate/opinion hook about relationship/gender dynamics — the strongest format across all 3 accounts (~40% US audience)",
    1, "", "Same hook text on a new Goth character reveal or mystery outfit transition (change ONLY the visual variable)",
]


def _used_range(sheet):
    max_row, max_col = 1, 1
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def _col_letter(n):
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


async def main():
    print("Lade altes Workbook...")
    raw = await sheets_tools.download_workbook()
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    print("Tabs im alten Workbook:", wb.sheetnames)

    service = sheets_tools._get_sheets_service()

    # Rename the new sheet's default tab to our first tab name, then add the rest.
    meta = service.spreadsheets().get(spreadsheetId=NEW_SPREADSHEET_ID).execute()
    default_sheet_id = meta["sheets"][0]["properties"]["sheetId"]

    requests = [{
        "updateSheetProperties": {
            "properties": {"sheetId": default_sheet_id, "title": TAB_ORDER[0]},
            "fields": "title",
        }
    }]
    for name in TAB_ORDER[1:]:
        requests.append({"addSheet": {"properties": {"title": name}}})

    print("Lege Tabs an...")
    service.spreadsheets().batchUpdate(spreadsheetId=NEW_SPREADSHEET_ID, body={"requests": requests}).execute()

    for name in TAB_ORDER:
        if name not in wb.sheetnames:
            print(f"UEBERSPRUNGEN (nicht im alten Workbook): {name}")
            continue
        sheet = wb[name]
        max_row, max_col = _used_range(sheet)
        values = []
        for r in range(1, max_row + 1):
            row_values = []
            for c in range(1, max_col + 1):
                cell = sheet.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    row_values.append("")
                elif hasattr(v, "isoformat"):  # datetime
                    row_values.append(v.strftime("%Y-%m-%d"))
                else:
                    row_values.append(v)
            values.append(row_values)

        range_str = f"'{name}'!A1:{_col_letter(max_col)}{max_row}"
        service.spreadsheets().values().update(
            spreadsheetId=NEW_SPREADSHEET_ID, range=range_str,
            valueInputOption="USER_ENTERED", body={"values": values},
        ).execute()
        print(f"Migriert: {name} ({max_row}x{max_col})")

    # Add the one real Scaling Log entry that never made it into any sheet.
    service.spreadsheets().values().append(
        spreadsheetId=NEW_SPREADSHEET_ID, range="'Scaling Log'!A1",
        valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS",
        body={"values": [MISSING_SCALING_LOG_ROW]},
    ).execute()
    print("Fehlenden echten Scaling-Log-Eintrag (lunaxvale/DbjcfCTh1ta) ergaenzt.")

    print("\nMigration fertig.")


if __name__ == "__main__":
    asyncio.run(main())
